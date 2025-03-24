import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime, timedelta
import time
import openpyxl

# Function to read parameters from Excel file
def read_parameters(file_path):
    xl = pd.ExcelFile(file_path)
    last_rows = {}
    for sheet_name in xl.sheet_names:
        if sheet_name.startswith('sample'):
            df = xl.parse(sheet_name)
            if not df.empty:
                last_row = df.iloc[-1]
                last_rows[sheet_name] = last_row.to_dict()
    return last_rows

# Function to convert seconds to HH:MM:SS format
def seconds_to_hms(seconds):
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    seconds = int(seconds % 60)
    return f"{hours:02}:{minutes:02}:{seconds:02}"

# Function to write results to Excel file
def write_results(file_path, results, sample_name):
    try:
        if os.path.exists(file_path):
            book = load_workbook(file_path)
        else:
            book = openpyxl.Workbook()
        sheet_name = sample_name
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            sheet = book.create_sheet(sheet_name)
            headers = list(results.keys())
            sheet.append(headers)
        new_row = list(results.values())
        sheet.append(new_row)
        book.save(file_path)
    except Exception as e:
        st.error(f"An error occurred: {e}")

# Function to calculate parameters with rounding
def helper_func_cal_param(params):
    end_of_drill_collar = round(params['end_of_drill_collar'], 2)
    d_hwdp_to_dp = round(params['d_hwdp_to_dp'], 2)
    d_drill_collar = round(params['d_drill_collar'], 2)
    d_riser = round(params['d_riser'], 2)
    surface_length = round(params['surface_length'], 2)
    current_hole_depth = round(params['current_hole_depth'], 2)
    pump_input_speed = round(params['pump_speed'], 2)
    last_casing_shoe_depth = round(params['last_casing_shoe_depth'], 2)
    internal_diameter_casing = round(params['internal_diameter_casing'], 2)
    diameter_open_hole = round(params['diameter_open_hole'], 2)
    sample_name = params['sample_name']

    len_of_casing_without_surface_length = round(last_casing_shoe_depth - surface_length, 2)
    casing_length = round(len_of_casing_without_surface_length + surface_length, 2)

    len_of_open_hole_from_casing_shoe = round(max(0, current_hole_depth - casing_length), 2)
    len_of_drill_pipe_in_open_hole = round(max(0, len_of_open_hole_from_casing_shoe - end_of_drill_collar), 2)
    len_of_drill_collar_in_casing = round(max(0, end_of_drill_collar - len_of_open_hole_from_casing_shoe), 2)
    len_of_drill_pipe_in_casing = round(len_of_casing_without_surface_length - len_of_drill_collar_in_casing, 2)
    len_of_drill_collar_in_open_hole = round(end_of_drill_collar - len_of_drill_collar_in_casing, 2)

    dia_open_hole_squared = round(diameter_open_hole**2, 2)
    d_drill_collar_squared = round(d_drill_collar**2, 2)
    d_hwdp_to_dp_squared = round(d_hwdp_to_dp**2, 2)
    internal_diameter_casing_squared = round(internal_diameter_casing**2, 2)
    d_riser_squared = round(d_riser**2, 2)

    k_constant = 0.000971
    pump_coe_constant = 0.1372

    annular_vol_at_open_hole = round(k_constant * (
        len_of_drill_collar_in_open_hole * (dia_open_hole_squared - d_drill_collar_squared) +
        len_of_drill_pipe_in_open_hole * (dia_open_hole_squared - d_hwdp_to_dp_squared)
    ), 2)
    annular_vol_at_casing_hole = round(k_constant * (
        len_of_drill_collar_in_casing * (internal_diameter_casing_squared - d_drill_collar_squared) +
        len_of_drill_pipe_in_casing * (internal_diameter_casing_squared - d_hwdp_to_dp_squared)
    ), 2)
    annular_vol_at_surface_wellhead_and_riser = round(k_constant * (
        surface_length * (d_riser_squared - d_hwdp_to_dp_squared)
    ), 2)

    pump_output_speed = round(pump_coe_constant * pump_input_speed, 2)
    estimated_lag_time_minutes = round((annular_vol_at_open_hole + annular_vol_at_casing_hole + annular_vol_at_surface_wellhead_and_riser) / pump_output_speed, 2) if pump_output_speed > 0 else 0
    estimated_lag_time_seconds = int(estimated_lag_time_minutes * 60) if pump_input_speed > 0 else 0
    estimated_lag_time_hms = seconds_to_hms(estimated_lag_time_seconds) if pump_input_speed > 0 else "read"

    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    results = {
        'sample_name': sample_name,
        'estimated_lag_time_seconds': estimated_lag_time_seconds,
        'estimated_lag_time_hms': estimated_lag_time_hms,
        'pump_output_speed': pump_output_speed,
        'annular_vol_at_surface_wellhead_and_riser': annular_vol_at_surface_wellhead_and_riser,
        'annular_vol_at_casing_hole': annular_vol_at_casing_hole,
        'annular_vol_at_open_hole': annular_vol_at_open_hole,
        'len_of_drill_collar_in_open_hole': len_of_drill_collar_in_open_hole,
        'len_of_drill_pipe_in_casing': len_of_drill_pipe_in_casing,
        'len_of_drill_pipe_in_open_hole': len_of_drill_pipe_in_open_hole,
        'len_of_drill_collar_in_casing': len_of_drill_collar_in_casing,
        'len_of_open_hole_from_casing_shoe': len_of_open_hole_from_casing_shoe,
        'new_cutting_depth': current_hole_depth,
        'is_at_surface': False,
        'published_datetime': current_datetime
    }
    return results

# Function to load completed samples from the output file
def load_completed_samples(output_file):
    completed_samples = set()
    if os.path.exists(output_file):
        xl = pd.ExcelFile(output_file)
        for sheet_name in xl.sheet_names:
            if sheet_name.startswith('sample'):
                df = xl.parse(sheet_name)
                if not df.empty and 'is_at_surface' in df.columns:
                    last_row = df.iloc[-1]
                    if last_row['is_at_surface']:
                        completed_samples.add(sheet_name)
    return completed_samples

# Function to load existing samples from output file
def load_existing_samples(output_file):
    existing_samples = {}
    if os.path.exists(output_file):
        xl = pd.ExcelFile(output_file)
        for sheet_name in xl.sheet_names:
            if sheet_name.startswith('sample'):
                df = xl.parse(sheet_name)
                if not df.empty:
                    last_row = df.iloc[-1].to_dict()
                    for key in ['estimated_lag_time_seconds', 'pump_output_speed', 'new_cutting_depth']:
                        if key in last_row and isinstance(last_row[key], (str, float, int)):
                            last_row[key] = float(last_row[key]) if isinstance(last_row[key], (str, int)) else last_row[key]
                    existing_samples[sheet_name] = last_row
    return existing_samples

# Helper function to convert image to base64 (required for embedding in HTML)
def get_base64_of_file(file_path):
    import base64
    with open(file_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode()
    
    
# Streamlit App
def main():
    # Custom CSS with green and orange theme and logo positioning
    st.markdown("""
        <style>
        /* Add padding to main content */
        .block-container {
            padding: 2rem !important;
            max-width: 100% !important;
        }
        /* Adjust title padding and set color to orange */
        h1 {
            padding-top: 2rem !important;
            color: #F28C38; /* Orange */
        }
        /* Subheaders in green */
        h2 {
            color: #2E7D32; /* Green */
        }
        /* Style columns */
        [data-testid="column"] {
            padding: 0.5rem !important;
        }
        /* Add demarcation line between columns */
        [data-testid="column"]:first-child {
            border-right: 1px solid #e0e0e0;
        }
        /* Style the countdown section with light green background */
        .countdown-section {
            background-color: #C8E6C9; /* Light Green */
            padding: 0.5rem;
            border: 1px solid #2E7D32; /* Green border */
        }
        /* Compact countdown table */
        .countdown-table td, .countdown-table th {
            padding: 5px !important;
            font-size: 14px !important;
        }
        /* Ensure main content takes remaining width */
        .main-content {
            padding: 0.5rem;
        }
        /* Sidebar header in green */
        .css-1v3fvcr h2 {
            color: #2E7D32; /* Green for sidebar header */
        }
        /* Logo positioning */
        .logo-container {
            position: absolute;
            top: 10px;
            right: 10px;
            z-index: 1000;
        }
        </style>
    """, unsafe_allow_html=True)

    # Add company logo in the top right
    script_dir = os.path.dirname(os.path.abspath(__file__))
    logo_path = os.path.join(script_dir, 'first-epdc-logo.png')  # Adjust this path to your logo file
    if os.path.exists(logo_path):
        # Use st.markdown to inject HTML with the image
        st.markdown(
            f"""
            <div class="logo-container">
                <img src="data:image/png;base64,{get_base64_of_file(logo_path)}" width="200">
            </div>
            """,
            unsafe_allow_html=True
        )
    else:
        st.warning("Logo file 'logo.png' not found in the script directory.")

    st.title("Drilling Lag Time Calculator")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, 'estimate_lag_time_input_data - Copy.xlsx')
    output_file = os.path.join(script_dir, 'estimate_lag_time_output_result - Copy.xlsx')

    if not os.path.exists(input_file):
        st.error(f"Input file '{input_file}' does not exist in the same folder as the script.")
        return

    current_time = datetime.now()

    if 'results_dict' not in st.session_state:
        st.session_state.results_dict = {}
        st.session_state.input_params = {}  # New dictionary to store original input parameters
        st.session_state.last_updated = {}
        st.session_state.remaining_times = {}
        st.session_state.completed_samples = load_completed_samples(output_file)
        st.session_state.previous_option = None
        
        existing_samples = load_existing_samples(output_file)
        for sample, results in existing_samples.items():
            st.session_state.results_dict[sample] = results
            st.session_state.last_updated[sample] = time.time()
            st.session_state.remaining_times[sample] = results['estimated_lag_time_seconds'] if not results['is_at_surface'] else 0
            # Since output file doesn't have input params, we'll need to initialize them later if needed

        inputs = read_parameters(input_file)
        if inputs:
            for sample, params in inputs.items():
                if sample in existing_samples:
                    continue
                for key, value in params.items():
                    if isinstance(value, (float, int)) and key != 'sample_name':
                        params[key] = round(value, 2) if isinstance(value, float) else value
                params['sample_name'] = sample
                results = helper_func_cal_param(params)
                st.session_state.results_dict[sample] = results
                st.session_state.input_params[sample] = params.copy()  # Store input params
                st.session_state.last_updated[sample] = time.time()
                st.session_state.remaining_times[sample] = results['estimated_lag_time_seconds']
                write_results(output_file, results, sample)
        else:
            st.warning("No data loaded from the input file.")

    st.sidebar.header("Options")
    option = st.sidebar.selectbox(
        "What would you like to do?",
        ("View Results", "Modify Pump Speed", "Update Parameters", "Add New Depth")
    )

    if option == "View Results" and st.session_state.previous_option != "View Results":
        st.session_state.previous_option = option
        st.rerun()

    if st.session_state.previous_option != option:
        st.session_state.previous_option = option

    content_container = st.container()

    with content_container:
        if option == "View Results":
            st.subheader("Current Results")
            
            if not st.session_state.results_dict and not st.session_state.completed_samples:
                st.warning("No results available to display.")
            else:
                table_data = []
                rerun_needed = False

                completed_data = {}
                if os.path.exists(output_file):
                    xl = pd.ExcelFile(output_file)
                    for sheet_name in xl.sheet_names:
                        if sheet_name.startswith('sample'):
                            df = xl.parse(sheet_name)
                            if not df.empty and 'is_at_surface' in df.columns:
                                last_row = df.iloc[-1].to_dict()
                                if last_row['is_at_surface'] and sheet_name not in st.session_state.results_dict:
                                    if isinstance(last_row['estimated_lag_time_seconds'], str):
                                        try:
                                            last_row['estimated_lag_time_seconds'] = int(float(last_row['estimated_lag_time_seconds']))
                                        except ValueError:
                                            last_row['estimated_lag_time_seconds'] = 0
                                    completed_data[sheet_name] = last_row

                for sample, results in st.session_state.results_dict.items():
                    lag_time_seconds = results['estimated_lag_time_seconds']
                    pump_speed = results['pump_output_speed'] / 0.1372
                    published_datetime_str = results['published_datetime']
                    is_at_surface = results['is_at_surface']

                    try:
                        published_datetime = datetime.strptime(published_datetime_str, "%Y-%m-%d %H:%M:%S")
                        elapsed_time = (current_time - published_datetime).total_seconds()
                    except ValueError:
                        elapsed_time = 0

                    if is_at_surface or sample in st.session_state.completed_samples:
                        remaining_time = 0
                    else:
                        remaining_time = max(0, int(lag_time_seconds - elapsed_time)) if pump_speed > 0 else 0

                    st.session_state.remaining_times[sample] = remaining_time

                    if remaining_time <= 0 and not is_at_surface and pump_speed > 0 and sample not in st.session_state.completed_samples:
                        st.session_state.completed_samples.add(sample)
                        st.write(f"Cuttings for depth {sample} have reached the surface.")
                        results_at_surface = results.copy()
                        results_at_surface['estimated_lag_time_seconds'] = 0
                        results_at_surface['estimated_lag_time_hms'] = '00:00:00'
                        results_at_surface['is_at_surface'] = True
                        results_at_surface['published_datetime'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        for key, value in results_at_surface.items():
                            if isinstance(value, float) and key != 'estimated_lag_time_seconds':
                                results_at_surface[key] = round(value, 2)
                        st.session_state.results_dict[sample] = results_at_surface
                        write_results(output_file, results_at_surface, sample)
                        st.rerun()

                    row_data = results.copy()
                    if pump_speed == 0 and not is_at_surface:
                        row_data['remaining_lag_time'] = "—"
                    else:
                        row_data['remaining_lag_time'] = "00:00:00" if (sample in st.session_state.completed_samples or is_at_surface) else ("read" if pump_speed == 0 else str(timedelta(seconds=remaining_time)))
                    row_data['is_completed'] = sample in st.session_state.completed_samples
                    table_data.append(row_data)

                    if remaining_time > 0 and pump_speed > 0 and not is_at_surface:
                        rerun_needed = True

                for sample, data in completed_data.items():
                    row_data = data.copy()
                    row_data['remaining_lag_time'] = "00:00:00"
                    row_data['is_completed'] = True
                    table_data.append(row_data)

                df = pd.DataFrame(table_data).round(2)
                df['estimated_lag_time_seconds'] = pd.to_numeric(df['estimated_lag_time_seconds'], errors='coerce').fillna(0).astype(int)

                priority_columns = ['sample_name', 'remaining_lag_time', 'new_cutting_depth', 'pump_output_speed', 'is_at_surface']
                other_columns = [col for col in df.columns if col not in priority_columns and col != 'is_completed']
                column_order = priority_columns + other_columns + ['is_completed']
                df = df[column_order]

                def style_df(val):
                    if isinstance(val, str) and val not in ["00:00:00", "read", "—"] and val.startswith('0'):
                        return 'background-color: #ffff99'
                    return ''

                def style_rows(row):
                    if row['is_completed'] or row['is_at_surface']:
                        return ['background-color: #e6ffe6'] * len(row)
                    elif row['pump_output_speed'] == 0:
                        return ['background-color: #ffcccc'] * len(row)
                    return [''] * len(row)

                styled_df = df.style.apply(style_rows, axis=1).map(style_df, subset=['remaining_lag_time'])
                final_df = styled_df.hide(axis='columns', subset=['is_completed'])

                st.dataframe(
                    final_df,
                    use_container_width=True,
                    height=(len(table_data) * 35 + 38) if len(table_data) < 10 else 400
                )

                if rerun_needed:
                    time.sleep(1)
                    st.rerun()

        elif option == "Modify Pump Speed":
            st.subheader("Modify Pump Speed for All Samples")
            new_speed = st.number_input("Enter new pump speed", min_value=0.0, step=0.1, format="%.2f")
            time_input = st.text_input("Enter time of modification (YYYY-MM-DD HH:MM:SS)", value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            if st.button("Update Pump Speed"):
                inputs = read_parameters(input_file)
                for sample, params in inputs.items():
                    if sample in st.session_state.completed_samples:
                        st.write(f"Skipping {sample} - already completed")
                        continue
                    
                    params['pump_speed'] = round(new_speed, 2)
                    params['time_at_which_change_occurred'] = time_input
                    
                    if sample in st.session_state.results_dict:
                        last_results = st.session_state.results_dict[sample]
                        init_lag_time = last_results['estimated_lag_time_seconds']
                        try:
                            last_time_published = datetime.strptime(last_results['published_datetime'], "%Y-%m-%d %H:%M:%S")
                            given_date = datetime.strptime(time_input, "%Y-%m-%d %H:%M:%S")
                            time_difference = given_date - last_time_published
                            difference_in_seconds = time_difference.total_seconds()
                            current_hole_depth = last_results['new_cutting_depth']
                            if init_lag_time > 0 and last_results['pump_output_speed'] > 0:
                                cutting_distance_travelled = round((current_hole_depth * difference_in_seconds) / init_lag_time, 2)
                                current_hole_depth = round(current_hole_depth - cutting_distance_travelled, 2)
                                params['current_hole_depth'] = current_hole_depth
                        except ValueError as e:
                            st.write(f"Error parsing dates for {sample}: {e}")

                    params['sample_name'] = sample
                    results = helper_func_cal_param(params)
                    results['published_datetime'] = time_input
                    st.session_state.results_dict[sample] = results
                    st.session_state.input_params[sample] = params.copy()  # Update input params
                    st.session_state.last_updated[sample] = time.time()
                    st.session_state.remaining_times[sample] = results['estimated_lag_time_seconds']
                    write_results(input_file, params, sample)
                    write_results(output_file, results, sample)
                st.success("Pump speed and cutting depth updated for active samples!")

        elif option == "Update Parameters":
            st.subheader("Update Parameters")
            st.write("This section is not fully implemented yet.")

        elif option == "Add New Depth":
            st.subheader("Add New Depth")
            sub_option = st.radio("Choose an option:", ("Copy Parameters, Input Depth & Time", "Input All Parameters"))
            
            if sub_option == "Copy Parameters, Input Depth & Time":
                new_depth = st.number_input("Enter new current hole depth", min_value=0.0, step=0.1, format="%.2f")
                time_input = st.text_input("Enter time of modification (YYYY-MM-DD HH:MM:SS)", value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                if st.button("Add Depth"):
                    if st.session_state.input_params:
                        # Find the sample with the latest last_updated timestamp
                        latest_sample = max(st.session_state.last_updated.items(), key=lambda x: x[1])[0]
                        template_params = st.session_state.input_params[latest_sample].copy()
                        pump_speed = template_params['pump_speed']
                    else:
                        # Fallback to input file if no input params exist
                        inputs = read_parameters(input_file)
                        if inputs:
                            _, template_params = next(iter(inputs.items()))
                            pump_speed = template_params['pump_speed']
                        else:
                            st.error("No existing samples to copy parameters from.")
                            return

                    new_params = template_params.copy()
                    new_params['current_hole_depth'] = round(new_depth, 2)
                    new_params['pump_speed'] = round(pump_speed, 2)  # Use the latest pump speed
                    new_params['time_at_which_change_occurred'] = time_input
                    new_params['sample_name'] = f'sample_{int(new_depth)}'
                    results = helper_func_cal_param(new_params)
                    sample = new_params['sample_name']
                    results['published_datetime'] = time_input
                    st.session_state.results_dict[sample] = results
                    st.session_state.input_params[sample] = new_params.copy()  # Store input params
                    st.session_state.last_updated[sample] = time.time()
                    st.session_state.remaining_times[sample] = results['estimated_lag_time_seconds']
                    write_results(input_file, new_params, sample)
                    write_results(output_file, results, sample)
                    st.success(f"New depth {sample} added with pump speed {new_params['pump_speed']}!")

            elif sub_option == "Input All Parameters":
                new_depth = st.number_input("Enter current hole depth", min_value=0.0, step=0.1, format="%.2f")
                pump_speed = st.number_input("Enter pump speed", min_value=0.0, step=0.1, format="%.2f")
                end_of_drill_collar = st.number_input("Enter end_of_drill_collar", min_value=0.0, step=0.1, format="%.2f")
                d_hwdp_to_dp = st.number_input("Enter d_hwdp_to_dp", min_value=0.0, step=0.1, format="%.2f")
                d_drill_collar = st.number_input("Enter d_drill_collar", min_value=0.0, step=0.1, format="%.2f")
                d_riser = st.number_input("Enter d_riser", min_value=0.0, step=0.1, format="%.2f")
                surface_length = st.number_input("Enter surface_length", min_value=0.0, step=0.1, format="%.2f")
                last_casing_shoe_depth = st.number_input("Enter last_casing_shoe_depth", min_value=0.0, step=0.1, format="%.2f")
                internal_diameter_casing = st.number_input("Enter internal_diameter_casing", min_value=0.0, step=0.1, format="%.2f")
                diameter_open_hole = st.number_input("Enter diameter_open_hole", min_value=0.0, step=0.1, format="%.2f")
                
                if st.button("Add Depth"):
                    new_params = {
                        'sample_name': f'sample_{int(new_depth)}',
                        'pump_speed': round(pump_speed, 2),
                        'current_hole_depth': round(new_depth, 2),
                        'end_of_drill_collar': round(end_of_drill_collar, 2),
                        'd_hwdp_to_dp': round(d_hwdp_to_dp, 2),
                        'd_drill_collar': round(d_drill_collar, 2),
                        'd_riser': round(d_riser, 2),
                        'surface_length': round(surface_length, 2),
                        'last_casing_shoe_depth': round(last_casing_shoe_depth, 2),
                        'internal_diameter_casing': round(internal_diameter_casing, 2),
                        'diameter_open_hole': round(diameter_open_hole, 2),
                        'time_at_which_change_occurred': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }
                    
                    results = helper_func_cal_param(new_params)
                    sample = new_params['sample_name']
                    results['published_datetime'] = new_params['time_at_which_change_occurred']
                    st.session_state.results_dict[sample] = results
                    st.session_state.input_params[sample] = new_params.copy()  # Store input params
                    st.session_state.last_updated[sample] = time.time()
                    st.session_state.remaining_times[sample] = results['estimated_lag_time_seconds']
                    write_results(input_file, new_params, sample)
                    write_results(output_file, results, sample)
                    st.success(f"New depth {sample} added!")

if __name__ == "__main__":
    main()